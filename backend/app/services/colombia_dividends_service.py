"""
Servicio de dividendos Colombia: boletín oficial BVC (Fechas Exdividendo) + precio TradingView.
Dividendos: boletín "Fechas Exdividendo" de BVC (CMS Hygraph -> XLSX oficial),
filtrado a los últimos 3 meses calendario completos (incluye el actual). Los
dividendos pagados en varias cuotas se separan en una fila por cuota (ver
_split_cuotas), cada una con su propio monto y fecha ex-dividendo.
Precio (solo para calcular yield): TradingView Overview — es lo único que se sigue
consultando en TV; el calendario de dividendos de TV (fetch_tv_colombia) quedó abajo comentado.
Moneda base: COP. Cada moneda que traiga el boletín se convierte a COP con su propia tasa FX.
"""
import calendar
import io
import os
import re
import datetime
import requests as http
import yfinance as yf
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
# from tradingview_scraper.symbols.cal import CalendarScraper  # ya no se usa: dividendos vienen del boletín oficial BVC
from tradingview_scraper.symbols.overview import Overview
from tradingview_scraper.symbols.symbol_markets import SymbolMarkets

SUPABASE_URL         = os.getenv("NEXT_PUBLIC_SUPABASE_URL")
SUPABASE_KEY         = os.getenv("NEXT_PUBLIC_SUPABASE_ANON_KEY")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")
TABLE                = "dividendos_colombia"
BATCH                = 50

# Whitelist de tickers permitidos para Colombia (BVC).
# BRK.B se normaliza a BRKB (sin punto) para coincidir con lo que devuelve TV.
# BVC agrega sufijo "CO" a algunos activos internacionales (MSFTCO, TSLACO...).
COLOMBIA_TICKER_WHITELIST: frozenset[str] = frozenset({
    "AGUASACO", "ALICORC1CO", "GOOGL", "AMZN", "AAPL", "PFAVAL",
    "BHI", "BOGOTA", "CHILECO", "BCICO", "BAC", "BRKB",
    "BVC", "CNEC", "CAPCO", "CELSIA", "CEMARGOS", "PFCEMARGOS",
    "CPACASC1CO", "CENCOSUDCO", "CENCOMALCO", "C", "COLBUNCO",
    "CCUCO", "BVNCO", "VAPORESCO", "CONCONCRET", "EIMI", "CSPX",
    "CORFICOLCF", "PFCORFICOL", "BAPCO", "PFDAVVNDA", "PFDAVIGRP",
    "SPXS", "AMDVASCCO", "ECOPETROL", "ANDINABCO", "ENTELCO",
    "CMPCCO", "COPECCO", "ENELAMCO", "ENELCHILCO", "ECLCO",
    "ENKA", "ETB", "ICHN", "EXITO", "FALABELLCO", "FERREYC1CO",
    "F", "GEHC", "GE", "COPX", "LIT", "GXTESCOL", "URA",
    "GRUPOARGOS", "PFGRUPOARG", "GRUPOAVAL", "GRUBOLIVAR",
    "CIBEST", "PFCIBEST", "GEB", "NUTRESA", "GRUPOSURA", "PFGRUPOSURA",
    "HCOLSEL", "HIVECO", "ICOLPCAP", "INRETC1CO", "IFSCO", "IAMCO",
    "IPCHBC1CO", "EQAC", "SGLD", "ISA", "IUIT", "IB01", "LQDA",
    "SDIA", "CBU7", "RBOT", "IBIT", "IJPA", "IWVL", "INRA",
    "D26ACO", "ID27CO", "D28ACO", "ID29CO", "ID30CO",
    "EMGA", "ISAC", "4BRZ", "IDSE", "SUAS", "I500CO",
    "IUES", "IUFS", "IUHC", "CFMITNIPCO", "ITAUCLCO", "JPEA",
    "JNJ", "JPM", "LTMCO", "META", "MSFTCO", "MINEROS",
    "NKE", "NUAMCO", "NU", "NVDA", "PARAUCOCO", "PEI", "PBR",
    "PFE", "MALLPLAZCO", "PROMIGAS", "QUINENCOCO", "RIPLEYCO",
    "BSANTANDCO", "SMUCO", "CVERDEC1CO", "PORT", "SQMBCO",
    "SCCOCO", "SUM", "TERPEL", "TSLACO", "KOCO", "JETS", "TIN",
    "GOAUCO", "UBER", "SDHA", "GDXCO", "SMHCO", "VOO",
    "CONCHATOCO", "VOLCABC1CO",
    "PFGRUPSURA",   # alias TV de PFGRUPOSURA (Grupo Sura preferencial)
    "ICOLCAP",      # alias TV de ICOLPCAP (ETF iColcap)
})

# Nombres completos por ticker (fuente: whitelist oficial)
NOMBRES_COLOMBIA: dict[str, str] = {
    "AGUASACO":    "Aguas Andinas S.A.",
    "ALICORC1CO":  "Alicorp S.A.",
    "GOOGL":       "Alphabet Inc.",
    "AMZN":        "Amazon",
    "AAPL":        "Apple Inc.",
    "PFAVAL":      "Aval Preferencial",
    "BHI":         "BAC Holding International",
    "BOGOTA":      "Banco de Bogotá",
    "CHILECO":     "Banco de Chile",
    "BCICO":       "Banco de Crédito e Inversiones",
    "BAC":         "Bank of America Corporation",
    "BRKB":        "Berkshire Hathaway Inc Class B",
    "BVC":         "Bolsa de Valores de Colombia",
    "CNEC":        "Canacol",
    "CAPCO":       "Cap S.A.",
    "CELSIA":      "Celsia",
    "CEMARGOS":    "Cementos Argos",
    "PFCEMARGOS":  "Cementos Argos Preferencial",
    "CPACASC1CO":  "Cementos Pacasmayo S.A.A.",
    "CENCOSUDCO":  "Cencosud SA",
    "CENCOMALCO":  "Cencosud Shopping S.A.",
    "C":           "Citigroup Inc.",
    "COLBUNCO":    "Colbún S.A.",
    "CCUCO":       "Compañía Cervecerías Unidas S.A.",
    "BVNCO":       "Compañía de Minas Buenaventura S.A.A.",
    "VAPORESCO":   "Compañía Sud Americana de Vapores S.A.",
    "CONCONCRET":  "Concreto",
    "EIMI":        "Core MSCI EM IMI UCITS ETF USD",
    "CSPX":        "Core S&P 500",
    "CORFICOLCF":  "Corficolombiana",
    "PFCORFICOL":  "Corficolombiana Preferencial",
    "BAPCO":       "Credicorp Limited",
    "PFDAVVNDA":   "Davivienda",
    "PFDAVIGRP":   "Davivienda Group S.A",
    "SPXS":        "Direxion Daily S&P 500 Bear 3X Shares ETF",
    "AMDVASCCO":   "DVA Silicon Fund",
    "ECOPETROL":   "Ecopetrol",
    "ANDINABCO":   "Embotelladora Andina S.A. Serie B",
    "ENTELCO":     "Empresa Nacional de Telecomunicaciones S.A.",
    "CMPCCO":      "Empresas CMPC S.A.",
    "COPECCO":     "Empresas Copec S.A.",
    "ENELAMCO":    "Enel Américas SA",
    "ENELCHILCO":  "Enel Chile S.A.",
    "ECLCO":       "Engie Energía Chile S.A.",
    "ENKA":        "Enka de Colombia S.A.",
    "ETB":         "Etb",
    "ICHN":        "iShares MSCI China UCITS ETF USD (Acc)",
    "EXITO":       "Éxito S.A.",
    "FALABELLCO":  "Falabella SA",
    "FERREYC1CO":  "Ferreycorp S.A.A.",
    "F":           "Ford",
    "GEHC":        "GE HealthCare Technologies Inc.",
    "GE":          "General Electric Company",
    "COPX":        "Global X Copper Miners ETF",
    "LIT":         "Global X Lithium & Battery Tech ETF",
    "GXTESCOL":    "Global X Tes Colombia ETF",
    "URA":         "Global X Uranium ETF",
    "GRUPOARGOS":  "Grupo Argos",
    "PFGRUPOARG":  "Grupo Argos Preferencial",
    "GRUPOAVAL":   "Grupo Aval",
    "GRUBOLIVAR":  "Grupo Bolívar",
    "CIBEST":      "Grupo Cibest",
    "PFCIBEST":    "Grupo Cibest Preferencial",
    "GEB":         "Grupo Energía de Bogotá",
    "NUTRESA":     "Grupo Nutresa",
    "GRUPOSURA":   "Grupo Suramericana",
    "PFGRUPOSURA": "Grupo Suramericana Preferencial",
    "PFGRUPSURA":  "Grupo Suramericana Preferencial",
    "HCOLSEL":     "Hcolsel",
    "HIVECO":      "HIVE Digital Technologies",
    "ICOLPCAP":    "Icolcap",
    "ICOLCAP":     "Icolcap",
    "INRETC1CO":   "InRetail Peru Corp",
    "IFSCO":       "Intercorp Financial Services Inc.",
    "IAMCO":       "Inversiones Aguas Metropolitanas S.A.",
    "IPCHBC1CO":   "Inversiones Portuarias Chancay S.A.A.",
    "EQAC":        "Invesco EQQQ Nasdaq-100 UCITS ETF",
    "SGLD":        "Invesco Physical Gold ETC",
    "ISA":         "Isa",
    "IUIT":        "iShare S&P 500 Tech",
    "IB01":        "iShare US Treas 0.1 YR USD A",
    "LQDA":        "iShares $ Corp Bond UCITS ETF",
    "SDIA":        "iShares $ Short Duration Corp Bond UCITS",
    "CBU7":        "iShares $ Treasury Bond 3-7yr UCITS ETF",
    "RBOT":        "iShares Automation & Robotics UCITS",
    "IBIT":        "iShares Bitcoin Trust",
    "IJPA":        "iShares Core MSCI Japan IMI UCITS ETF USD Acc",
    "IWVL":        "iShares Edge MSCI World Value Fact UCITS",
    "INRA":        "iShares Global Clean Energy UCITS ETF USD (Acc)",
    "D26ACO":      "iShares iBonds Dec 2026 Term $ Corp UCITS ETF",
    "ID27CO":      "iShares iBonds Dec 2027 Term $ Corp UCITS ETF",
    "D28ACO":      "iShares iBonds Dec 2028 Term $ Corp UCITS ETF",
    "ID29CO":      "iShares iBonds Dec 2029 Term $ Corp UCITS ETF",
    "ID30CO":      "iShares iBonds Dec 2030 Term $ Corp UCITS ETF",
    "EMGA":        "iShares J.P. Morgan EM Local Govt Bond UCITS ETF",
    "ISAC":        "iShares MSCI ACWI",
    "4BRZ":        "iShares MSCI Brazil UCITS ETF",
    "IDSE":        "iShares MSCI Europe SRI UCITS ETF USD (Acc)",
    "SUAS":        "iShares MSCI USA SRI",
    "I500CO":      "iShares S&P 500 Colombia",
    "IUES":        "iShares S&P 500 Energy Sector UCITS (Acc)",
    "IUFS":        "iShares S&P 500 Financials",
    "IUHC":        "iShares SP 500 Health Care Sector UCITS (Acc)",
    "CFMITNIPCO":  "IT NOW S&P IPSA",
    "ITAUCLCO":    "Itaú Corpbanca",
    "JPEA":        "J. P. Morgan USD EM Bond",
    "JNJ":         "Johnson & Johnson",
    "JPM":         "JPMorgan Chase & Co",
    "LTMCO":       "Latam",
    "META":        "Meta Platforms, Inc.",
    "MSFTCO":      "Microsoft Corp",
    "MINEROS":     "Mineros",
    "NKE":         "Nike, Inc.",
    "NUAMCO":      "Nuam Exchange",
    "NU":          "Nubank",
    "NVDA":        "Nvidia Corporation",
    "PARAUCOCO":   "Parque Arauco S.A.",
    "PEI":         "Patrimonio Autónomo Estrategias Inmobiliarias",
    "PBR":         "Petróleo Brasileiro S.A",
    "PFE":         "Pfizer",
    "MALLPLAZCO":  "Plaza S.A.",
    "PROMIGAS":    "Promigas",
    "QUINENCOCO":  "Quiñenco S.A.",
    "RIPLEYCO":    "Ripley Corporación S.A.",
    "BSANTANDCO":  "Santander Chile Holding S.A.",
    "SMUCO":       "SMU S.A.",
    "CVERDEC1CO":  "Sociedad Minera Cerro Verde S.A.",
    "PORT":        "Sociedad Portafolio",
    "SQMBCO":      "Sociedad Química y Minera de Chile",
    "SCCOCO":      "Southern Copper Corp",
    "SUM":         "Summit Materials",
    "TERPEL":      "Terpel",
    "TSLACO":      "Tesla, Inc.",
    "KOCO":        "The Coca-Cola Company",
    "JETS":        "The U.S. Global Jets ETF",
    "TIN":         "Títulos Inmobiliarios TIN",
    "GOAUCO":      "U.S. Global GO GOLD",
    "UBER":        "Uber",
    "SDHA":        "USD Corp Bond UCITS ETF",
    "GDXCO":       "VanEck Gold Miners ETF",
    "SMHCO":       "VanEck Semiconductor ETF",
    "VOO":         "Vanguard 500 Index Fund ETF",
    "CONCHATOCO":  "Viña Concha y Toro S.A.",
    "VOLCABC1CO":  "Volcan Compañía Minera S.A.A.",
}


def _get_nombre_colombia(ticker: str) -> str | None:
    """Nombre desde dict estático. Normaliza punto (BRK.B→BRKB).
    Si el ticker termina en 'CO', prueba también sin el sufijo."""
    t = ticker.strip().upper().replace(".", "")
    if t in NOMBRES_COLOMBIA:
        return NOMBRES_COLOMBIA[t]
    if len(t) > 2 and t.endswith("CO") and t[:-2] in NOMBRES_COLOMBIA:
        return NOMBRES_COLOMBIA[t[:-2]]
    return None


# ── Helpers ───────────────────────────────────────────────────────────────────
def _in_whitelist(ticker: str) -> bool:
    """Verifica si un ticker está en la whitelist de Colombia.
    Normaliza puntos (BRK.B -> BRKB) y maneja el sufijo CO que BVC agrega
    a algunos activos internacionales (GOOGLCO -> GOOGL como red de seguridad).
    """
    t = ticker.strip().upper().replace(".", "")
    if t in COLOMBIA_TICKER_WHITELIST:
        return True
    if len(t) > 2 and t.endswith("CO") and t[:-2] in COLOMBIA_TICKER_WHITELIST:
        return True
    return False


def _ts_to_date(ts) -> str | None:
    if not ts:
        return None
    return datetime.datetime.fromtimestamp(ts, tz=datetime.timezone.utc).strftime("%Y-%m-%d")


def _to_cop(amount, currency, tc) -> float | None:
    if amount is None:
        return None
    return round(amount * tc, 2) if currency == "USD" else round(amount, 2)


def get_tc_usdcop() -> float:
    return round(yf.Ticker("USDCOP=X").fast_info["last_price"], 2)


# ── Precios BVC via TradingView Overview ──────────────────────────────────────
_US_EXCHANGE_PRIORITY = ("NYSE", "NASDAQ", "AMEX")


def _resolve_mgc_symbol(ticker: str) -> str | None:
    """Referencias del Mercado Global Colombiano (ticker extranjero + sufijo 'CO',
    ej. GECO->GE, JPMCO->JPM) no tienen feed propio bajo BVC: en TradingView —
    son solo el reenvío de la bolsa de origen. Se resuelve el símbolo real
    buscando el ticker sin el sufijo y filtrando por coincidencia exacta."""
    if not (len(ticker) > 2 and ticker.endswith("CO")):
        return None
    stripped = ticker[:-2]

    try:
        candidates = SymbolMarkets().scrape(symbol=stripped).get("data") or []
    except Exception:
        return None

    matches = [
        c for c in candidates
        if c.get("type") == "stock" and c.get("symbol", "").split(":")[-1].upper() == stripped.upper()
    ]
    if not matches:
        return None

    for exch in _US_EXCHANGE_PRIORITY:
        for c in matches:
            if c.get("exchange") == exch:
                return c["symbol"]

    matches.sort(key=lambda c: c.get("market_cap_basic") or 0, reverse=True)
    return matches[0]["symbol"]


def _get_bvc_prices(symbols: list[str]) -> dict[str, dict]:
    """Precio actual para tickers de Colombia. Intenta primero BVC:{sym} (precio en COP);
    si no existe (referencias MGC como GECO/JPMCO), resuelve el símbolo real en su bolsa
    de origen (NYSE/NASDAQ/...), cuyo precio viene en la moneda nativa de esa bolsa, no en COP.
    Se marca `is_cop` para que el cálculo de yield use el monto en la moneda correcta."""
    if not symbols:
        return {}

    overview = Overview()

    def _price_of(symbol: str) -> float | None:
        try:
            result = overview.get_symbol_overview(symbol=symbol)
            if result.get("status") == "success":
                price = result["data"].get("close")
                if price:
                    return float(price)
        except Exception:
            pass
        return None

    def _fetch(sym: str):
        price = _price_of(f"BVC:{sym}")
        if price is not None:
            return sym, {"price": price, "is_cop": True}

        resolved = _resolve_mgc_symbol(sym)
        if resolved:
            price = _price_of(resolved)
            if price is not None:
                return sym, {"price": price, "is_cop": False}

        return sym, None

    with ThreadPoolExecutor(max_workers=8) as ex:
        return dict(ex.map(_fetch, symbols))


# ── Tasa de cambio {moneda}->COP via TradingView Overview (FX_IDC) ────────────
def _get_fx_rate_to_cop(currency: str, cache: dict) -> float | None:
    """Tasa {currency}->COP. Cachea por moneda dentro de la misma ejecución."""
    currency = (currency or "COP").strip().upper()
    if currency == "COP":
        return 1.0
    if currency in cache:
        return cache[currency]

    rate = None
    try:
        overview = Overview()
        result = overview.get_symbol_overview(symbol=f"FX_IDC:{currency}COP")
        if result.get("status") == "success":
            close = result["data"].get("close")
            if close:
                rate = round(float(close), 4)
    except Exception:
        rate = None

    cache[currency] = rate
    return rate


# ── Fetch TradingView Colombia (LEGACY — reemplazado por boletín oficial BVC) ─
# Se deja comentado en vez de borrarse por si hay que revertir. El calendario de
# dividendos ahora viene de fetch_bvc_colombia(); TV solo se sigue usando arriba
# en _get_bvc_prices()/_get_fx_rate_to_cop() para precio y tasa de cambio.
#
# def fetch_tv_colombia(tc: float) -> list[dict]:
#     now   = datetime.datetime.now(tz=datetime.timezone.utc)
#     today = now.replace(hour=0, minute=0, second=0, microsecond=0)
#     ts_from = int((today - datetime.timedelta(weeks=4)).timestamp())
#     ts_to   = int((today + datetime.timedelta(weeks=8, seconds=86399)).timestamp())
#
#     raw = CalendarScraper().scrape_dividends(
#         timestamp_from=ts_from,
#         timestamp_to=ts_to,
#         markets=["colombia"],
#     )
#
#     seen: dict[tuple, dict] = {}
#     for ev in raw:
#         ex_date = _ts_to_date(ev.get("dividend_ex_date_upcoming")) or \
#                   _ts_to_date(ev.get("dividend_ex_date_recent"))
#         if not ex_date:
#             continue
#         full_sym = ev.get("full_symbol", "")
#         ticker   = full_sym.split(":")[-1]
#         if not _in_whitelist(ticker):
#             continue
#         key = (full_sym, ex_date)
#         if key in seen:
#             continue
#         amount   = ev.get("dividend_amount_upcoming") or ev.get("dividend_amount_recent")
#         currency = ev.get("fundamental_currency_code", "USD")
#         monto_cop = _to_cop(amount, currency, tc)
#         seen[key] = {
#             "symbol"         : full_sym,
#             "nombre"         : _get_nombre_colombia(ticker) or ev.get("name") or ev.get("description"),
#             "fuente"         : "TV",
#             "fecha_corte"    : ex_date,
#             "fecha_pago"     : _ts_to_date(ev.get("dividend_payment_date_upcoming")) or
#                                _ts_to_date(ev.get("dividend_payment_date_recent")),
#             "monto_original" : amount,
#             "moneda_original": currency,
#             "tc_usdcop"      : tc if currency == "USD" else None,
#             "monto_cop"      : monto_cop,
#             "tipo"           : "efectivo",
#             "en_partes"      : False,
#             "concepto"       : None,
#             "yield_tv_pct"   : ev.get("dividends_yield"),
#         }
#     rows = list(seen.values())
#
#     # Para los sin yield en TV, calcular con precio COP de Overview
#     sin_yield = [
#         r["symbol"].split(":")[-1]
#         for r in rows
#         if r["yield_tv_pct"] is None and r["monto_cop"]
#     ]
#     if sin_yield:
#         precios = _get_bvc_prices(sin_yield)
#         for r in rows:
#             if r["yield_tv_pct"] is None and r["monto_cop"]:
#                 sym_short = r["symbol"].split(":")[-1]
#                 precio = precios.get(sym_short)
#                 if precio and precio > 0:
#                     r["yield_tv_pct"] = round(r["monto_cop"] / precio * 100, 6)
#
#     return rows


# ── Boletín oficial BVC: "Fechas Exdividendo" (CMS Hygraph -> XLSX) ───────────
BVC_HYGRAPH_URL = "https://us-east-1-bolsa-co.cdn.hygraph.com/content/ckdolgg6k07rc01xnc22d25r1/master"
BVC_EXDIVIDEND_CATEGORY_ID = "ckgp9xu6g2ddw098110jscm0f"

# Token público de solo-lectura embebido en el bundle JS del sitio de BVC (no es
# secreto: se envía al navegador de cualquier visitante). Se usa como fast-path;
# si BVC lo rota, _discover_hygraph_token() lo vuelve a extraer del sitio en
# caliente, así que nunca hay que actualizarlo a mano. El valor de arranque
# puede overridearse por env si se quiere fijar uno propio.
_hygraph_token_cache: dict[str, str | None] = {
    "token": os.getenv(
        "BVC_HYGRAPH_TOKEN",
        "eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCIsImtpZCI6ImdjbXMtbWFpbi1wcm9kdWN0aW9uIn0."
        "eyJ2ZXJzaW9uIjozLCJpYXQiOjE2Mzk2ODQ5NDUsImF1ZCI6WyJodHRwczovL2FwaS11cy1lYXN0LTEtYm9sc2EtY28uZ3JhcGhjbXMuY29tL3YyL2NrZG9sZ2c2azA3cmMwMXhuYzIyZDI1cjEvbWFzdGVyIiwiaHR0cHM6Ly9tYW5hZ2VtZW50LW5leHQuZ3JhcGhjbXMuY29tIl0sImlzcyI6Imh0dHBzOi8vbWFuYWdlbWVudC5ncmFwaGNtcy5jb20vIiwic3ViIjoiMmNmMzY5ZTMtNWYwMS00YjUwLThhNGMtZTM3YmQxNDI4ZGYwIiwianRpIjoiY2tpcDdjaWhxODdsNDAxejEwcHplOTI4byJ9."
        "ZFDpgyyGMnav1J6BRVt_ZT43vLFzQgf-iWddP2BmzgjCm1-zx_qZFMIRMo0q7TRqDgFWhpQIt2Xuku0ens8KfvdkaPpnkeqzeZYufXeIjpXgyF_nF0tXsZjN3eSm-LkLYQK65dcJsA1UwJJjUu9wk1i-sDjnBU2LiXhFc6XCMXoH912cG3eb9sCwWatodMrkrV4qgK-zdKU2nc3FGJwL2X-lUtwdvnmKhRmLbnAzTmi4pLlog7KWYof5Syk44ysYF3stThRb8uJA580Wgcw7WEFOSJerpvdvnGPxHNYudSJdxQ2kF4SchJUnohUNdrSkmAgkogtWy-Vs1uB7Px03b4QhjjxYMTnocoTY_0a7r1pulvi0vtp_foOD-XgPlR1qoza6g9LW1KLJ39gBVv6SD_TLZ4j94HvR3iloXN8iIta_KGS4zXHg10ay6ZI1SXvomSbHxaOzZXclk1yWZDv76n1ZbwDWDfRkgf4figWLKqJCenL-_uIdXBe3lnD8odBLDqvQ_BM-We2MrCAPWzGC2Vo_fIpyYYDsS8h4AUZw3m3E7w1j_61fFJOIIRDpmSA49KA9Dy4hlO-eLQVvTPUWw-Tn6QAfZDrc-slr_EgjTMJg0KCP4e3NYlDPxWmbQKUzErZj_GwDp85ezaCJTFapCy1A2Ax1CVirdk88TbjoqJE",
    )
}

_APP_CHUNK_RE = re.compile(r'/_next/static/chunks/pages/_app-[a-zA-Z0-9]+\.js')
_TOKEN_NEAR_RE = re.compile(r'cdn\.hygraph\.com[^"]*"[^}]*?content:"([^"]+)"')


def _discover_hygraph_token() -> str | None:
    """Extrae en caliente el token público de Hygraph desde el chunk _app de bvc.com.co
    (se carga en toda la web, no solo en /informes-y-boletines). Self-healing: si BVC
    rota el token, esto lo vuelve a encontrar sin tocar código ni variables de entorno."""
    try:
        home = http.get(
            "https://www.bvc.com.co/",
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
            timeout=15,
        )
        home.raise_for_status()
        chunk_match = _APP_CHUNK_RE.search(home.text)
        if not chunk_match:
            return None

        chunk = http.get(
            f"https://www.bvc.com.co{chunk_match.group(0)}",
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Referer": "https://www.bvc.com.co/",
            },
            timeout=15,
        )
        chunk.raise_for_status()
        token_match = _TOKEN_NEAR_RE.search(chunk.text)
        return token_match.group(1) if token_match else None
    except Exception:
        return None


def _hygraph_query(query: str, variables: dict, operation_name: str) -> dict:
    """POST a Hygraph con el token cacheado. Si falla por auth (401 o sin data),
    redescubre el token desde el sitio y reintenta una vez antes de fallar."""
    payload = {"query": query, "variables": variables, "operationName": operation_name}

    def _post(token: str | None):
        return http.post(
            BVC_HYGRAPH_URL,
            json=payload,
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            timeout=20,
        )

    r = _post(_hygraph_token_cache["token"])
    auth_failed = r.status_code in (401, 403) or "errors" in r.json()
    if auth_failed:
        fresh_token = _discover_hygraph_token()
        if fresh_token and fresh_token != _hygraph_token_cache["token"]:
            _hygraph_token_cache["token"] = fresh_token
            r = _post(fresh_token)

    r.raise_for_status()
    body = r.json()
    if "errors" in body:
        raise RuntimeError(f"Hygraph (BVC) rechazó la consulta: {body['errors']}")
    return body["data"]


def _fetch_bvc_bulletin_meta() -> dict:
    """Consulta el CMS de BVC y devuelve el boletín vigente de Fechas Exdividendo (fileName + url)."""
    query = (
        "query ReportsAndBulletinsByCategoryAndDate($categoryId: ID!, $locales: [Locale!]!) {"
        "  pdReportsAndBulletins(where: {category: {id: $categoryId}}, stage: PUBLISHED, "
        "    locales: $locales, orderBy: date_DESC, first: 1) {"
        "    date title attached { fileName url }"
        "  }"
        "}"
    )
    data = _hygraph_query(
        query,
        {"categoryId": BVC_EXDIVIDEND_CATEGORY_ID, "locales": ["es_CO"]},
        "ReportsAndBulletinsByCategoryAndDate",
    )
    reports = data["pdReportsAndBulletins"]
    if not reports:
        raise RuntimeError("BVC no tiene publicado el boletín de Fechas Exdividendo")
    return reports[0]["attached"]


def _download_bvc_bulletin() -> bytes:
    meta = _fetch_bvc_bulletin_meta()
    r = http.get(meta["url"], timeout=30)
    r.raise_for_status()
    return r.content


def _norm_header(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split()).strip().upper()


def _cell_to_date(value) -> str | None:
    """Convierte celda de fecha (datetime o serial de Excel) a 'YYYY-MM-DD'."""
    if value is None or value == "" or value == "-":
        return None
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return (_EXCEL_EPOCH + datetime.timedelta(days=float(value))).isoformat()
        except (OverflowError, ValueError):
            return None
    return None


def _find_bvc_header_row(ws) -> tuple[int, dict[str, int]]:
    """Ubica la fila de headers en español (contiene NEMOTECNICO) y mapea columna->índice.
    El layout de columnas varía por año, por eso se busca por nombre y no por índice fijo."""
    for row in ws.iter_rows(min_row=1, max_row=15):
        values = [_norm_header(c.value) for c in row]
        if "NEMOTECNICO" not in values:
            continue
        col: dict[str, int] = {}
        for idx, v in enumerate(values):
            if v == "EMISOR":
                col["emisor"] = idx
            elif v == "NEMOTECNICO":
                col["ticker"] = idx
            elif v == "MODO DE PAGO":
                col["modo_pago"] = idx
            elif v == "MONEDA":
                col["moneda"] = idx
            elif v in {"FECHA INICIAL EX-DIVIDENDO", "FECHA INICIAL"}:
                col["fecha_ex"] = idx
            elif v in {"FECHA FINAL EX-DIVIDENDO", "FECHA FINAL Y DE PAGO", "FECHA FINAL"}:
                col["fecha_ex_final"] = idx
            elif v == "FECHA DE PAGO":
                col["fecha_pago"] = idx
            elif v == "VALOR TOTAL DEL DIVIDENDO":
                col["valor_total"] = idx
            elif v == "VALOR CUOTA":
                col["valor_cuota"] = idx
            elif v == "DESCRIPCION PAGO PDU":
                col["descripcion"] = idx
        return row[0].row, col
    raise RuntimeError("No se encontró la fila de headers (NEMOTECNICO) en el boletín BVC")


# ── Dividendos en varias cuotas (ej. "PAGADEROS EN DOS CUOTAS 29JUL Y 18DIC") ──
# La fila del boletín solo trae la cuota MÁS PRÓXIMA (fecha ex-dividendo real,
# VALOR CUOTA real). Las demás cuotas solo aparecen descritas en texto libre, con
# su fecha de PAGO (no de ex-dividendo). Por eso la fecha ex-dividendo de las
# cuotas 2+ es una estimación (mismo offset ex->pago que la cuota real conocida).
_MESES_ES = {
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12,
}
_NUM_CUOTAS_ES = {
    "UNA": 1, "UNO": 1, "DOS": 2, "TRES": 3, "CUATRO": 4, "CINCO": 5, "SEIS": 6, "SIETE": 7,
}
_N_CUOTAS_RE = re.compile(r'(\d{1,2}|UNA|UNO|DOS|TRES|CUATRO|CINCO|SEIS|SIETE)\s+CUOTAS?', re.IGNORECASE)
_MES_ABR = "ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC"
# Entre el monto y la fecha SIEMPRE debe haber moneda y/o "EL" — si no, "29JUL" se
# puede partir por backtracking en monto="2" + día="9" (bug real, ya visto con GEB).
_MONTO_FECHA_RE = re.compile(
    rf'([\d]+(?:[.,]\d+)?)\s*(?:(?:COP|USD|CLP|PEN)\s*(?:EL\s*)?|EL\s*)(\d{{1,2}})\s*(?:DE\s+)?({_MES_ABR})\s*/?\s*(\d{{2,4}})?',
    re.IGNORECASE,
)
# "EL" es opcional aquí porque BVC a veces escribe "29JUL Y 18DIC" sin "EL" (ej. GEB);
# sin un monto por delante no hay ambigüedad de backtracking como en el regex de arriba.
# "DE" opcional cubre "16 DE DICIEMBRE" (nombre de mes completo, no solo abreviado).
_FECHA_SOLA_RE = re.compile(rf'(?:EL\s*)?(\d{{1,2}})\s*(?:DE\s+)?({_MES_ABR})\s*/?\s*(\d{{2,4}})?', re.IGNORECASE)


def _resolve_cuota_year(month: int, ref_year: int, ref_month: int, explicit_year: str | None) -> int:
    if explicit_year:
        y = int(explicit_year)
        return 2000 + y if y < 100 else y
    return ref_year + 1 if month < ref_month else ref_year


def _dates_from_matches(matches: list[tuple], ref_date: datetime.date) -> list[datetime.date]:
    out = []
    ref_year, ref_month = ref_date.year, ref_date.month
    for m in matches:
        day = int(m[-3])
        month = _MESES_ES[m[-2].upper()]
        year = _resolve_cuota_year(month, ref_year, ref_month, m[-1] or None)
        out.append(datetime.date(year, month, day))
        ref_year, ref_month = year, month
    return out


def _split_cuotas(fecha_ex: str, fecha_pago: str | None, valor_total: float, valor_cuota: float,
                   descripcion: str | None) -> list[dict]:
    """Si la descripción menciona varias cuotas con sus propias fechas de pago,
    devuelve una entrada por cuota (fecha_ex real para la primera, estimada para
    las siguientes; monto de cada cuota, no el total). Si no se puede separar con
    confianza (formato de texto no reconocido, o reparto ambiguo con 3+ cuotas sin
    monto propio), devuelve una sola fila con VALOR CUOTA — nunca con el total."""
    base = [{"fecha_ex": fecha_ex, "fecha_pago": fecha_pago, "monto": valor_cuota}]
    if not descripcion:
        return base

    n_match = _N_CUOTAS_RE.search(descripcion)
    if not n_match:
        return base
    raw_n = n_match.group(1).upper()
    n_cuotas = int(raw_n) if raw_n.isdigit() else _NUM_CUOTAS_ES.get(raw_n)
    if not n_cuotas or n_cuotas < 2:
        return base

    ex_dt = datetime.date.fromisoformat(fecha_ex)
    pago_dt = datetime.date.fromisoformat(fecha_pago) if fecha_pago else ex_dt
    offset_days = (pago_dt - ex_dt).days

    monto_matches = _MONTO_FECHA_RE.findall(descripcion)
    if len(monto_matches) == n_cuotas:
        pay_dates = _dates_from_matches(monto_matches, ex_dt)
        montos = [float(m[0].replace(",", ".")) for m in monto_matches]
    else:
        fecha_matches = _FECHA_SOLA_RE.findall(descripcion)
        if len(fecha_matches) != n_cuotas:
            return base  # texto no reconocido: no se arriesga a repartir mal
        pay_dates = _dates_from_matches(fecha_matches, ex_dt)
        reparto_igual = "IGUAL" in descripcion.upper() or (
            valor_total and abs(valor_total / n_cuotas - valor_cuota) < max(0.01, valor_cuota * 0.01)
        )
        if reparto_igual:
            montos = [round(valor_total / n_cuotas, 4)] * n_cuotas
        elif n_cuotas == 2:
            montos = [valor_cuota, round(valor_total - valor_cuota, 4)]
        else:
            return base  # 3+ cuotas desiguales sin monto propio: no hay dato confiable

    rows = [{"fecha_ex": fecha_ex, "fecha_pago": fecha_pago, "monto": valor_cuota}]
    for pay_dt, monto in list(zip(pay_dates, montos))[1:]:
        ex_i = pay_dt - datetime.timedelta(days=offset_days)
        rows.append({"fecha_ex": ex_i.isoformat(), "fecha_pago": pay_dt.isoformat(), "monto": monto})
    return rows


def _month_range(end_year: int, end_month: int, n_months: int) -> tuple[datetime.date, datetime.date]:
    """Rango de n_months meses calendario completos, terminando en end_year-end_month."""
    end_date = datetime.date(end_year, end_month, calendar.monthrange(end_year, end_month)[1])
    start_index = (end_year * 12 + (end_month - 1)) - (n_months - 1)
    start_year, start_month0 = divmod(start_index, 12)
    start_date = datetime.date(start_year, start_month0 + 1, 1)
    return start_date, end_date


def _parse_bvc_sheet(ws, header_row: int, col: dict, start_date: datetime.date, end_date: datetime.date) -> list[dict]:
    """Recorre todas las filas de la hoja (no solo las ancladas al rango: la fecha
    ex-dividendo REAL de la fila puede quedar "vieja" — ej. la primera cuota de un
    dividendo de todo el año — aunque otra de sus cuotas sí caiga en el rango).
    Regla: se separan las cuotas primero; si AL MENOS UNA cae en el rango, se
    devuelven TODAS las cuotas de esa fila juntas, para no partir un dividendo
    a la mitad. Si ninguna cuota cae en el rango, la fila completa se descarta."""
    fecha_pago_idx = col.get("fecha_pago", col.get("fecha_ex_final"))
    rows: list[dict] = []
    # +2: la hoja tiene header en español seguido de header en inglés antes de los datos
    for row in ws.iter_rows(min_row=header_row + 2):
        values = [c.value for c in row]

        ticker = values[col["ticker"]] if col["ticker"] < len(values) else None
        if not ticker or not str(ticker).strip():
            continue
        ticker = str(ticker).strip().upper()
        if not _in_whitelist(ticker):
            continue

        fecha_ex = _cell_to_date(values[col["fecha_ex"]]) if col["fecha_ex"] < len(values) else None
        if not fecha_ex:
            continue

        valor_total = values[col["valor_total"]] if col.get("valor_total") is not None and col["valor_total"] < len(values) else None
        valor_cuota = values[col["valor_cuota"]] if col.get("valor_cuota") is not None and col["valor_cuota"] < len(values) else None
        try:
            valor_total = float(valor_total) if valor_total is not None else None
            valor_cuota = float(valor_cuota) if valor_cuota is not None else valor_total
        except (TypeError, ValueError):
            continue
        if valor_cuota is None:
            continue

        emisor = values[col["emisor"]] if col.get("emisor") is not None and col["emisor"] < len(values) else None
        moneda = values[col["moneda"]] if col.get("moneda") is not None and col["moneda"] < len(values) else "COP"
        modo_pago = values[col["modo_pago"]] if col.get("modo_pago") is not None and col["modo_pago"] < len(values) else None
        descripcion = values[col["descripcion"]] if col.get("descripcion") is not None and col["descripcion"] < len(values) else None
        descripcion = str(descripcion).strip() if descripcion else None
        fecha_pago = _cell_to_date(values[fecha_pago_idx]) if fecha_pago_idx is not None and fecha_pago_idx < len(values) else None

        cuotas = _split_cuotas(fecha_ex, fecha_pago, valor_total or valor_cuota, valor_cuota, descripcion)
        any_in_range = any(
            start_date <= datetime.date.fromisoformat(c["fecha_ex"]) <= end_date for c in cuotas
        )
        if not any_in_range:
            continue

        for cuota in cuotas:
            rows.append({
                "ticker"     : ticker,
                "emisor"     : str(emisor).strip() if emisor else None,
                "moneda"     : str(moneda).strip().upper() if moneda else "COP",
                "modo_pago"  : str(modo_pago).strip() if modo_pago else None,
                "fecha_ex"   : cuota["fecha_ex"],
                "fecha_pago" : cuota["fecha_pago"],
                "monto"      : cuota["monto"],
                "descripcion": descripcion,
            })
    return rows


def _parse_bvc_range(xlsx_bytes: bytes, start_date: datetime.date, end_date: datetime.date) -> list[dict]:
    """Lee las hojas de año que toque el rango (puede cruzar diciembre->enero) y
    devuelve las cuotas (ver _split_cuotas) cuya fecha ex-dividendo cae en el rango."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    rows: list[dict] = []
    for year in range(start_date.year, end_date.year + 1):
        sheet_name = f"FECHAS EX-DIVIDENDO {year}"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row, col = _find_bvc_header_row(ws)
        if "ticker" not in col or "fecha_ex" not in col:
            wb.close()
            raise RuntimeError(f"Headers inesperados en la hoja '{sheet_name}' del boletín BVC")
        rows.extend(_parse_bvc_sheet(ws, header_row, col, start_date, end_date))

    wb.close()
    return rows


def fetch_bvc_colombia(n_months: int = 3) -> list[dict]:
    """Dividendos de los últimos n_months meses calendario (incluye el actual)
    desde el boletín oficial BVC. El yield se calcula con precio de TradingView
    Overview (único uso restante de TV)."""
    now = datetime.datetime.now(tz=datetime.timezone.utc)
    start_date, end_date = _month_range(now.year, now.month, n_months)
    xlsx_bytes = _download_bvc_bulletin()
    raw_rows = _parse_bvc_range(xlsx_bytes, start_date, end_date)

    fx_cache: dict[str, float | None] = {}
    seen: dict[tuple, dict] = {}
    for r in raw_rows:
        tc = _get_fx_rate_to_cop(r["moneda"], fx_cache)
        monto_cop = round(r["monto"] * tc, 2) if tc is not None else None
        full_sym = f"BVC:{r['ticker']}"
        key = (full_sym, r["fecha_ex"])
        if key in seen:
            continue
        seen[key] = {
            "symbol"         : full_sym,
            "nombre"         : _get_nombre_colombia(r["ticker"]) or r["emisor"],
            "fuente"         : "BVC",
            "fecha_corte"    : r["fecha_ex"],
            "fecha_pago"     : r["fecha_pago"],
            "monto_original" : r["monto"],
            "moneda_original": r["moneda"],
            "tc_usdcop"      : tc if r["moneda"] != "COP" else None,
            "monto_cop"      : monto_cop,
            "tipo"           : "efectivo" if r["modo_pago"] and "EFECTIVO" in r["modo_pago"].upper() else (r["modo_pago"] or "efectivo"),
            "en_partes"      : False,
            "concepto"       : r["descripcion"],
            "yield_tv_pct"   : None,
        }
    rows = list(seen.values())

    # Yield vía precio de TradingView Overview (BVC no publica precio/yield en el boletín)
    tickers = [r["symbol"].split(":")[-1] for r in rows if r["monto_cop"]]
    if tickers:
        precios = _get_bvc_prices(tickers)
        for r in rows:
            if not r["monto_cop"]:
                continue
            info = precios.get(r["symbol"].split(":")[-1])
            if not info or not info["price"] or info["price"] <= 0:
                continue
            # BVC:{ticker} da precio en COP -> comparar contra monto_cop.
            # Resuelto en bolsa extranjera (NYSE/NASDAQ/...) -> precio en moneda
            # nativa de esa bolsa, que es la misma que moneda_original del boletín.
            numerador = r["monto_cop"] if info["is_cop"] else r["monto_original"]
            r["yield_tv_pct"] = round(numerador / info["price"] * 100, 6)

    return rows


# ── Preview ───────────────────────────────────────────────────────────────────
def get_preview() -> dict:
    now = datetime.datetime.now(tz=datetime.timezone.utc)
    start_date, end_date = _month_range(now.year, now.month, 3)
    rows = fetch_bvc_colombia()

    return {
        "fecha_preview"    : now.strftime("%Y-%m-%d %H:%M UTC"),
        "rango_consultado" : {"desde": start_date.isoformat(), "hasta": end_date.isoformat()},
        "resumen": {
            "bvc_total" : len(rows),
            "con_yield" : sum(1 for r in rows if r["yield_tv_pct"] is not None),
            "sin_yield" : sum(1 for r in rows if r["yield_tv_pct"] is None),
        },
        "dividendos": rows,
    }


# ── Sync a Supabase ───────────────────────────────────────────────────────────
def _supabase_headers():
    key = SUPABASE_SERVICE_KEY or SUPABASE_KEY
    return {
        "apikey"       : key,
        "Authorization": f"Bearer {key}",
        "Content-Type" : "application/json",
        "Prefer"       : "resolution=merge-duplicates,return=minimal",
    }


def sync_to_supabase(rows: list[dict]) -> dict:
    if not SUPABASE_URL or not (SUPABASE_SERVICE_KEY or SUPABASE_KEY):
        raise RuntimeError("Faltan credenciales Supabase en .env")

    url     = f"{SUPABASE_URL}/rest/v1/{TABLE}"
    results = {"guardados": 0, "errores": []}

    for i in range(0, len(rows), BATCH):
        batch = rows[i:i + BATCH]
        r = http.post(url, json=batch, headers=_supabase_headers(), timeout=30)
        if r.status_code in (200, 201, 204):
            results["guardados"] += len(batch)
        else:
            results["errores"].append(
                f"batch {i//BATCH+1}: {r.status_code} {r.text[:100]}"
            )
    return results
